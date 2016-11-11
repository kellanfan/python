import os, commands, socket, time, threading, re
from openpyxl.workbook import Workbook 
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter
#from openpyxl import load_workbook
import psycopg2

#for exel
wb = Workbook()
ew = ExcelWriter(workbook = wb)
dest_filename = r'result.xlsx'
ws_instance = wb.worksheets[0]
ws_instance.title = "instance"

ws_LB = wb.create_sheet()
ws_LB.title = "LB"

conn = psycopg2.connect(database="zone", user="postgres", password="zhu88jie", host="192.168.180.132", port="5432")
cur_instance = conn.cursor()
cur_instance.execute("select instance.owner, image.base_id, router_static.val1, router_static.val2, instance.instance_name from (router_static LEFT OUTER JOIN instance on router_static.val1 = instance.instance_id) left join image on instance.image_id = image.image_id where router_static.val1 like '%i-%';")
rows_instance = cur_instance.fetchall()
print len(rows_instance)
lies_instance = len(rows_instance[0])

cur_LB = conn.cursor()
cur_LB.execute("select owner from loadbalancer where status = 'active' order by 1;")
rows_LB = cur_LB.fetchall()
print len(rows_LB)
lies_LB = len(rows_LB[0])
print lies_LB

def data(row_num,rows_name):
    datainfo = []
    for i in rows_name:
        datainfo.append(i[row_num])
    f = datainfo
    return f


def intertexel(column_num, f, sheet_name):
    count = len(f)
    for n in range(1, count):
        sheet_name.cell(row=n, column=column_num).value = '%s' %(f[n])



for lie in range(1, lies_instance):
    f = []
    f = data(lie, rows_instance)
    intertexel(lie, f, ws_instance)

for lie in range(1, lies_LB):
    f = []
    f = data(lie, rows_LB)
    intertexel(lie, f, ws_LB)


conn.commit()
cur_instance.close()
cur_LB.close()
conn.close()
ew.save('result.xlsx')
