import os, commands, socket, time, threading, re
#import openpyxl
from openpyxl.workbook import Workbook 
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter



wb = Workbook()
ew = ExcelWriter(workbook = wb)
dest_filename = r'result.xlsx'
ws_test2 = wb.worksheets[0]
ws_test2.title = "test2"
f = open('./userid')
lines = f.readlines()
count = len(lines)
for n in range(1, count):
    ws_test2.cell(row=n, column=1).value = '%s' %(lines[n])
        
ew.save('result.xlsx')
